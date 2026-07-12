"""
services/report_service.py
============================
Generates CSV, Excel, JSON and PDF reports from a list of `ScanResult`.

Each report row includes: Image Name, Barcode, OCR, Similarity,
Confidence, Status, Processing Time, Reason, Timestamp.
"""

from __future__ import annotations

import io
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.models import ScanResult
from logger import get_logger

log = get_logger(__name__)

COLUMNS = [
    "Image Number", "File Name", "Barcode", "Barcode Type", "Barcode Engine",
    "OCR Text", "OCR Confidence", "OCR Engine",
    "Similarity %", "Match Confidence %", "Status", "Reason",
    "Processing Time", "Scan Time",
]


def _to_row(r: ScanResult) -> Dict[str, Any]:
    flat = r.to_flat_dict()
    return {
        "Image Number": flat["image_number"],
        "File Name": flat["filename"],
        "Barcode": flat["barcode_value"],
        "Barcode Type": flat["barcode_type"],
        "Barcode Engine": flat["barcode_engine"],
        "OCR Text": flat["ocr_text"],
        "OCR Confidence": f"{flat['ocr_confidence']:.1f}%",
        "OCR Engine": flat["ocr_engine"],
        "Similarity %": f"{flat['similarity_percent']:.1f}%",
        "Match Confidence %": f"{flat['match_confidence']:.1f}%",
        "Status": flat["match_status"],
        "Reason": flat["mismatch_reason"],
        "Processing Time": f"{flat['processing_time']:.3f}s",
        "Scan Time": flat["scan_time"],
    }


class ReportService:
    """Persists and exports scan results across CSV, Excel, JSON and PDF formats."""

    def __init__(self, output_dir: str = "outputs") -> None:
        self.output_dir = output_dir
        self.csv_path = os.path.join(output_dir, "results.csv")
        self.xlsx_path = os.path.join(output_dir, "results.xlsx")
        self.json_dir = os.path.join(output_dir, "json")
        self.reports_dir = os.path.join(output_dir, "reports")
        Path(self.json_dir).mkdir(parents=True, exist_ok=True)
        Path(self.reports_dir).mkdir(parents=True, exist_ok=True)

    # ── DataFrame ────────────────────────────────────────────────────────

    def to_dataframe(self, results: List[ScanResult]) -> pd.DataFrame:
        return pd.DataFrame([_to_row(r) for r in results], columns=COLUMNS)

    # ── Disk persistence ─────────────────────────────────────────────────

    def save_to_csv(self, results: List[ScanResult]) -> bool:
        try:
            self.to_dataframe(results).to_csv(self.csv_path, index=False)
            log.info(f"CSV saved -> {self.csv_path}")
            return True
        except Exception as exc:
            log.error(f"save_to_csv error: {exc}")
            return False

    def save_to_excel(self, results: List[ScanResult]) -> bool:
        try:
            rows = [_to_row(r) for r in results]
            wb = Workbook()
            ws = wb.active
            ws.title = "Barcode OCR Results"

            hdr_fill = PatternFill("solid", fgColor="2563EB")
            hdr_font = Font(color="FFFFFF", bold=True, size=11)
            center = Alignment(horizontal="center", vertical="center")

            for ci, col in enumerate(COLUMNS, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, center

            green_fill, green_font = PatternFill("solid", fgColor="D1FAE5"), Font(color="065F46", bold=True)
            red_fill, red_font = PatternFill("solid", fgColor="FEE2E2"), Font(color="991B1B", bold=True)
            amber_fill, amber_font = PatternFill("solid", fgColor="FEF3C7"), Font(color="92400E", bold=True)

            for ri, row in enumerate(rows, 2):
                status = row.get("Status", "")
                for ci, col in enumerate(COLUMNS, 1):
                    cell = ws.cell(row=ri, column=ci, value=row[col])
                    cell.alignment = center
                    if col == "Status":
                        if status == "MATCH":
                            cell.fill, cell.font = green_fill, green_font
                        elif status == "MISMATCH":
                            cell.fill, cell.font = red_fill, red_font
                        else:
                            cell.fill, cell.font = amber_fill, amber_font

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max(14, max_len + 4)

            wb.save(self.xlsx_path)
            log.info(f"Excel saved -> {self.xlsx_path}")
            return True
        except Exception as exc:
            log.error(f"save_to_excel error: {exc}")
            return False

    def save_all_json(self, results: List[ScanResult]) -> bool:
        try:
            path = os.path.join(self.json_dir, "all_results.json")
            with open(path, "w", encoding="utf-8") as fh:
                json.dump([r.to_flat_dict() for r in results], fh, indent=2, default=str)
            log.info(f"JSON saved -> {path}")
            return True
        except Exception as exc:
            log.error(f"save_all_json error: {exc}")
            return False

    def save_to_pdf(self, results: List[ScanResult], summary: Dict[str, Any] | None = None) -> str | None:
        """Generate a printable PDF summary report. Returns the file path, or None on failure."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(self.reports_dir, f"report_{ts}.pdf")
            doc = SimpleDocTemplate(path, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = [Paragraph("Barcode + OCR Verification Report", styles["Title"]), Spacer(1, 10)]

            if summary:
                summary_lines = " | ".join(f"{k}: {v}" for k, v in summary.items())
                elements.append(Paragraph(summary_lines, styles["Normal"]))
                elements.append(Spacer(1, 14))

            rows = [_to_row(r) for r in results]
            table_cols = ["Image Number", "File Name", "Barcode", "OCR Text", "Similarity %", "Status", "Reason"]
            data = [table_cols] + [[row.get(c, "") for c in table_cols] for row in rows]

            table = Table(data, repeatRows=1)
            style = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
            for ri, row in enumerate(rows, 1):
                if row.get("Status") == "MATCH":
                    style.append(("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#D1FAE5")))
                elif row.get("Status") == "MISMATCH":
                    style.append(("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#FEE2E2")))
            table.setStyle(TableStyle(style))
            elements.append(table)

            doc.build(elements)
            log.info(f"PDF report saved -> {path}")
            return path
        except Exception as exc:
            log.error(f"save_to_pdf error: {exc}")
            return None

    def load_results(self) -> List[Dict[str, Any]]:
        try:
            if os.path.isfile(self.csv_path):
                df = pd.read_csv(self.csv_path, dtype=str)
                return df.fillna("").to_dict("records")
            return []
        except Exception as exc:
            log.error(f"load_results error: {exc}")
            return []

    # ── In-memory bytes for Streamlit download buttons ──────────────────

    def get_csv_bytes(self, results: List[ScanResult]) -> bytes:
        return self.to_dataframe(results).to_csv(index=False).encode("utf-8")

    def get_excel_bytes(self, results: List[ScanResult]) -> bytes:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            self.to_dataframe(results).to_excel(writer, index=False, sheet_name="Barcode OCR Results")
        return buf.getvalue()

    def get_json_bytes(self, results: List[ScanResult]) -> bytes:
        return json.dumps([r.to_flat_dict() for r in results], indent=2, default=str).encode("utf-8")

    def get_pdf_bytes(self, results: List[ScanResult], summary: Dict[str, Any] | None = None) -> bytes | None:
        path = self.save_to_pdf(results, summary)
        if path is None:
            return None
        with open(path, "rb") as fh:
            return fh.read()
